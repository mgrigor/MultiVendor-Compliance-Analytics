import os
import sys
import time
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy

start = time.ctime()
print('''
################################################################################
#                                                                              #
#                   Samsung CELL Part 2 Exception Audit v1.0                   #
#                                                                              #
################################################################################
''')

directory = 'Raw Data'
tech = 'CELL'
level = 'Part 2'


def Select_subMarket():
    """Prompts the user to select a valid regional submarket folder."""
    print('Available subMarkets:')
    while True:
        items = os.listdir('Raw Data')
        valid_folders = [item for item in items if os.path.isdir(os.path.join('Raw Data', item))]
        for folder in valid_folders:
            print(f" ↳ {folder}")
        choice = input("\nPlease select subMarket from the list: ").upper()
        if choice in valid_folders:
            return choice
        print("\n❌ subMarket not found, please try again.")


def Samsung_Parameters(tech, level):
    """Dynamically loads required parameter metrics targets from the master blueprint sheet."""
    print(f'\n > Generating dataframe headers for required parameters ({tech} {level})...')
#    file = 'Samsung Required Parameters Master.xlsx'   # for live network
    file = 'Demo_GPL_Samsung.xlsx'  # For demo purposes
    
    df = pd.read_excel(file, sheet_name='master', index_col=False, dtype=str)

    cell_group = df.loc[(df['Technology'] == tech) & (df['Group'] == level)]
    
    reqd_parameters = []
    group_script_name_dict = {}
    groupby_script_name = cell_group.groupby('Group Script Name')
    
    for group_script, data in groupby_script_name:
        parameters = [item.split('[')[0] for item in data['Parameter'].unique().tolist()]
        parameters = sorted(list(set(parameters)))
        reqd_parameters.append(parameters)
        group_script_name_dict.update({tuple(parameters): group_script})

    if tech.upper() == 'ENB':
        labels = ['date', 'subMarket', 'market', 'enbname', 'enbid']
    else:
        labels = ['date', 'subMarket', 'market', 'enbname', 'enbid', 'cellId', 'sector', 'carrier', 'carrierId']
    
    headers_and_values = {label: '' for label in labels}
    param_mappings = dict(zip(cell_group['Parameter'], cell_group['Airwave Setting']))
    headers_and_values.update(param_mappings)
    
    df_headers = pd.DataFrame(headers_and_values, index=[0])
    df_headers = df_headers.map(lambda x: x.lower().strip() if isinstance(x, str) else x)
    
    return (reqd_parameters, group_script_name_dict, df_headers)


def load_data_files():
    """Loads transformed cell log files, executes Rule 1/2 controls, and cuts non-blueprint weight."""
    print(f'\n > Loading data files and executing Rule 1 & 2 structural controls...\n')
    path = os.path.join(directory, subMarket, tech, level)
    if not os.path.exists(path):
        path = os.path.join(directory, subMarket, tech)
        
    if not os.path.exists(path):
        print(f" [!] Error: Staging directory not found -> {path}")
        sys.exit()
        
    files = [file for file in os.listdir(path) if '-x.csv' in file]
    
    frames = []
    common_set = None
    
    for file in files:
        df = pd.read_csv(os.path.join(path, file), index_col=False, sep=',', dtype=str)
        df['enbid'] = df['enbid'].astype(str).str.zfill(6)
        
        SubMarket = df['subMarket'].dropna().unique().tolist()[0].split('-')[-1]
        print(f"   Staging SubMarket: {SubMarket} | File: {file}")
        
        if subMarket != SubMarket:
            print('   [!] Invalid regional tracking submarket detected! Aborting execution.')
            sys.exit()

        for label in ['cellId', 'sector', 'carrier']:
            df[label] = df[label].fillna('0').astype(float).astype(int).astype(str)

        # ─── OPTIMIZATION STEP A: CONSTRUCT CARRIERID FIRST AS MAIN ANCHOR ───
        df['carrierId'] = df['enbid'] + '-' + df['cellId'] + '-' + df['sector'] + '-' + df['carrier']
        
        # ─── OPTIMIZATION STEP B: DROP UNTRACKED PARAMETER WEIGHT IMMEDIATELY ───
        df = df.drop(columns=['bar-manual'], errors='ignore')
        
        columns = df.columns.tolist()
        columns.remove('carrierId')
        columns.insert(8, 'carrierId')
        df = df[columns]

        # ─── IMPLEMENTATION OF RULE 1: FRONT-DOOR "GROW" DATA PURGE ───
        init_rows = df.shape[0]
        df = df[~df['enbname'].str.lower().str.startswith('grow', na=False)].reset_index(drop=True)
        grow_purged = init_rows - df.shape[0]
        if grow_purged > 0:
            print(f"      [+] Purged {grow_purged} non-operational GROW rows from memory footprint.")

        # ─── IMPLEMENTATION OF RULE 2: VECTORIZED 4G5G MODERNIZATION PRIORITY ───
        df['is_modernized'] = df['enbname'].str.contains('4G5G', na=False, case=False)
        df.sort_values(by=['carrierId', 'is_modernized'], ascending=[True, False], inplace=True)
        # Slices duplicate rows keeping 'first' (forces active 4G5G data to overwrite legacy records)
        df.drop_duplicates(subset=['carrierId'], keep='first', inplace=True)
        df.drop(columns=['is_modernized'], inplace=True)

        current_file_set = set(df['carrierId'].unique())
        if common_set is None:
            common_set = current_file_set
        else:
            common_set = common_set.intersection(current_file_set)
        
        frames.append(df)

    print(f'\n > Filtering dataframes based on common carrier IDs...')
    common_carrierIDs = list(common_set) if common_set is not None else []
    
    filtered_frames = []
    for frame in frames:
        filtered_frame = frame[frame['carrierId'].isin(common_carrierIDs)]
        print(f"   ↳ Dimension Profile Matrix: {filtered_frame.shape}")
        filtered_frames.append(filtered_frame)
        
    return (filtered_frames, common_carrierIDs)


def checking_multiple_values_in_a_row(filtered_frames, common_carrierIDs):
    """Detects and flattens remaining multi-value records (e.g., symmetric parameters)."""
    print(f'\n > Running final compression checks on carrier layers...')
    cleaned_frames = []
    dup_carrierIDs = []
    
    for df in filtered_frames:
        columns = df.columns.tolist()
        meta_data_cols = columns[:9]
        parameter_cols = columns[9:]
        
        is_dup = df.duplicated(subset=['carrierId'], keep=False)
        
        if is_dup.any():
            clean_df = df[~is_dup]
            dup_df = df[is_dup]
            print(f'   Optimizing {dup_df["carrierId"].nunique()} nested multi-layer carrier elements...')
            dup_carrierIds = dup_df['carrierId'].unique().tolist()
            for carrierId in dup_carrierIds:
                if carrierId not in dup_carrierIDs:
                    dup_carrierIDs.append(carrierId)
                    
            agg_dict = {col: 'first' for col in meta_data_cols if col not in ['carrierId', 'enbname']}
            
            def agg_enbname(series):
                unique_vals = series.dropna().unique()
                if len(unique_vals) > 1:
                    for val in unique_vals:
                        if '4g5g' in str(val).lower():
                            return str(val)
                    return str(unique_vals[0])
                return str(unique_vals[0]) if len(unique_vals) > 0 else ""
                
            agg_dict['enbname'] = agg_enbname
            
            def agg_parameters(series):
                unique_vals = [str(x) for x in series.dropna().unique()]
                if len(unique_vals) > 1:
                    return '/'.join(unique_vals)
                return unique_vals[0] if len(unique_vals) > 0 else ""
                
            for col in parameter_cols:
                agg_dict[col] = agg_parameters
            
            compressed_df = dup_df.groupby('carrierId', as_index=False).agg(agg_dict)
            compressed_df = compressed_df[columns]
            df_final = pd.concat([clean_df, compressed_df], ignore_index=True)
            cleaned_frames.append(df_final)
        else:
            cleaned_frames.append(df)
            
    return (cleaned_frames, dup_carrierIDs)


def merging_dataframes(cleaned_frames):
    """Executes side-by-side horizontal outer merges on sector index keys and filters CBRS cells."""
    print('\n > Merging dataframes into a consolidated horizontal baseline...')
    if not cleaned_frames:
        return None, ""
    path = directory+'/'+subMarket
    cbrs_file = [f for f in os.listdir(path) if 'CBRS.xlsx' in f][0]
    cbrs_file = path+'/'+cbrs_file
    print(cbrs_file)
    
    if os.path.exists(cbrs_file):
        print(f'   > Reading CBRS cell exclusions from {cbrs_file}...')
        cbrs_df = pd.read_excel(cbrs_file, dtype=str)
        cbrs_col = [col for col in cbrs_df.columns if 'carrierid' in col.lower()]
        cbrs_cells = set(cbrs_df[cbrs_col[0]].dropna().unique()) if cbrs_col else set(cbrs_df.iloc[:, 0].dropna().unique())
    else:
        print(f'   ⚠️ Warning: {cbrs_file} not found! Proceeding without CBRS filtering.')
        cbrs_cells = set()

    merged_df = cleaned_frames[0]
    if len(cleaned_frames) > 1:
        for frame in cleaned_frames[1:]:
            parameters = frame.columns.tolist()[8:]
            merged_df = pd.merge(merged_df, frame[parameters], on='carrierId', how='outer')
    
    print('   > Removing CBRS cells...')
    merged_df = merged_df[~merged_df['carrierId'].isin(cbrs_cells)]

    headers = df_headers.columns.tolist()
    existing_headers = [col for col in headers if col in merged_df.columns]
    merged_df = merged_df[existing_headers]
       
    print('   > Removing duplicate rows...')
    merged_df = merged_df.drop_duplicates()

    print('   > Checking for carrierId[s] with multiple rows...')
    is_dup = merged_df.duplicated(subset=['carrierId'], keep=False)

    if is_dup.any():
        clean_rows = merged_df[~is_dup]
        duplicate_rows = merged_df[is_dup]
        
        def pick_clean_name(series):
            names = [str(n) for n in series.dropna().unique() if '/' not in str(n)]
            for n in names:
                if '4g5g' in n.lower():
                    return n
            return names[0] if names else str(series.iloc[0])
            
        resolved_names = duplicate_rows.groupby('carrierId')['enbname'].agg(pick_clean_name)
        compressed_rows = duplicate_rows.astype(str).groupby('carrierId', as_index=False).agg(
            lambda x: '/'.join(sorted(list(x.unique())))
        )
        compressed_rows['enbname'] = compressed_rows['carrierId'].map(resolved_names)
        compressed_rows = compressed_rows[clean_rows.columns]
        merged_df_clean = pd.concat([clean_rows, compressed_rows], ignore_index=True)
    else:
        print('     No multi-row carrierIds found.')
        merged_df_clean = merged_df
        
    print('   > Appending engineering target baseline values to row 0...')
    subMarket_report_new = pd.concat([df_headers, merged_df_clean], ignore_index=True)

    filename = os.path.join(directory, subMarket, f"Samsung_{subMarket}_CELL_Part_2.xlsx")
    return (subMarket_report_new, filename)    


def ac_barring_mo_signaling_usage_feature(subMarket_report_new1):
    """Conditionally prunes dead 'skip' parameters if the main switch is disabled."""
    print('\n > Analyzing ac-barring-mo-signaling-usage feature status...')
    df = subMarket_report_new1.copy()
    columns = df.columns.tolist()

    feature = 'ac-barring-mo-signaling-usage'
    feature_parameters = [column for column in columns if column.startswith(feature)]

    if not feature_parameters:
        return df

    data = df.iloc[1:]
    unique_states = pd.unique(data[feature_parameters].values.ravel())
    feature_state_list = sorted([str(x) for x in unique_states if pd.notna(x)])

    feature_state = '/'.join(feature_state_list) if feature_state_list else 'unknown'
    print(f'     Feature status across market: {feature_state}')
    
    if feature_state == 'disable':
        print('   > Pruning dead parameters associated with disabled feature switch...')
        skip_parameters = [column for column in columns if 'skip' in column]
        if skip_parameters:
            df = df.drop(columns=skip_parameters)
    return df


def get_common_enodebs_markets_and_carrierIds():
    """ADAPTIVE LOOKUP LAYER: Dynamically extracts common profiles across all three baseline reports."""
    print(f'\n > Filtering out common enodebs/markets/carrierIds across aligned layers...')    
    
    path_market = os.path.join(directory, subMarket)
    files = os.listdir(path_market)
    raw_dfs = {}
    
    for file in files:
        if file.startswith('~$') or not file.endswith('.xlsx'):
            continue
        
        file_lower = file.lower()
        if 'enb' in file_lower and 'exception' not in file_lower and 'report' not in file_lower:
            key = 'ENB'
        elif 'part' in file_lower and '1' in file_lower and 'exception' not in file_lower and 'report' not in file_lower:
            key = 'Part 1'
        elif 'part' in file_lower and '2' in file_lower and 'exception' not in file_lower and 'report' not in file_lower:
            key = 'Part 2'
        else:
            continue

        full_path = os.path.join(path_market, file)
        print(f"     ✅ Successfully Located {key} Ledger -> {file}")
        df = pd.read_excel(full_path, index_col=False, dtype=str)
        df = df.fillna(' ') 
        df = df.map(lambda x: x.lower().strip() if isinstance(x, str) else x)
        raw_dfs[key] = {'path': full_path, 'df': df}

    if len(raw_dfs) < 3:
        print("\n ❌ CRITICAL BLOCKED: Missing one or more required baseline sheets (ENB, Part 1, or Part 2)!")
        sys.exit()

    enb_data = raw_dfs['ENB']['df'].iloc[1:]
    p1_data = raw_dfs['Part 1']['df'].iloc[1:]
    p2_data = raw_dfs['Part 2']['df'].iloc[1:]

    valid_markets = sorted(list(set(enb_data['market']) & set(p1_data['market']) & set(p2_data['market'])))
    common_enodebs = sorted(list(set(enb_data['enbid']) & set(p1_data['enbid']) & set(p2_data['enbid'])))
    valid_carrierIds = set(p1_data['carrierId']) & set(p2_data['carrierId'])
    
    print(f'     Total Overlapping Nodes    : {len(common_enodebs)}')       
    print(f'     Total Overlapping Antennas : {len(valid_carrierIds)}')
    
    new_dfs = {}
    for key, file_info in raw_dfs.items():
        df = file_info['df']
        headers = df.iloc[0:1]
        data = df.iloc[1:]
        
        filtered_data = data[data['market'].isin(valid_markets) & data['enbid'].isin(common_enodebs)]
        if key in ['Part 1', 'Part 2']:
            filtered_data = filtered_data[filtered_data['carrierId'].isin(valid_carrierIds)]
        
        final_df = pd.concat([headers, filtered_data], ignore_index=True)
        new_filename = file_info['path'].replace('.xlsx', ' Report.xlsx')
        new_dfs[new_filename] = final_df
        save_to_excel(new_filename, final_df, report_type='ENB' if key == 'ENB' else 'CELL')
        
    common_items = [common_enodebs, valid_markets, list(valid_carrierIds)]
    return (common_items, new_dfs)


def retain_only_noncompliant_data(df, report_type='CELL', report_name=""):
    """Isolates data rows where configurations mismatch the blueprint baseline rules."""
    print(f'\n > Isolating non-compliant rows for {report_name}...')
    if len(df) <= 1:
        return df
        
    meta_count = 5 if report_type.upper() == 'ENB' else 9
    param_cols = df.columns[meta_count:]
    
    baseline_row = df.iloc[[0]]
    baseline_params = df.iloc[0][param_cols]
    live_data = df.iloc[1:]
    
    is_auditable_param = (baseline_params != '') & (baseline_params != ' ') & (baseline_params.notna())
    active_baseline = baseline_params[is_auditable_param]
    
    mismatches_dict = {}
    for col in active_baseline.index:
        allowed_vals = [v.strip() for v in str(active_baseline[col]).split(',') if v.strip()]
        
        def check_cell_mismatch(live_val):
            live_parts = [p.strip() for p in str(live_val).split('/') if p.strip()]
            if not live_parts:
                return False 
            return not any(part in allowed_vals for part in live_parts)
            
        mismatches_dict[col] = live_data[col].apply(check_cell_mismatch)
        
    mismatches = pd.DataFrame(mismatches_dict, index=live_data.index)
    non_compliant_mask = mismatches.any(axis=1)
    non_compliant_data = live_data[non_compliant_mask]
    
    return pd.concat([baseline_row, non_compliant_data], ignore_index=True)


def remove_compliant_columns(df, report_type='CELL'):
    """Prunes out columns where all running nodes meet engineering target guidelines."""
    print(f'\n > Pruning fully compliant parameter columns for {report_type}...')
    if len(df) <= 1:
        return df

    meta_count = 5 if report_type.upper() == 'ENB' else 9
    meta_cols = df.columns[:meta_count].tolist()
    param_cols = df.columns[meta_count:].tolist()
    
    baseline_row = df.iloc[0]
    live_data = df.iloc[1:]
    
    columns_to_keep = list(meta_cols)
    dropped_count = 0
    
    for col in param_cols:
        baseline_str = str(baseline_row[col]).strip()
        if baseline_str in ['', ' ', 'nan']:
            columns_to_keep.append(col)
            continue
            
        allowed_vals = [v.strip() for v in baseline_str.split(',') if v.strip()]
        
        def is_cell_compliant(live_val):
            live_parts = [p.strip() for p in str(live_val).split('/') if p.strip()]
            if not live_parts:
                return True
            return any(part in allowed_vals for part in live_parts)
            
        is_column_compliant = live_data[col].apply(is_cell_compliant).all()
        
        if is_column_compliant:
            dropped_count += 1
        else:
            columns_to_keep.append(col)
            
    print(f'   [+] Pruned Parameters   : {dropped_count}')
    print(f'   [+] Retained Deviations : {len(columns_to_keep) - meta_count}')
    return df[columns_to_keep]


def save_to_excel(filename, df, report_type='CELL', is_exception=False):
    """Exports processed dataframes to Excel via xlsxwriter and highlights exceptions."""
    df = df.copy().fillna(' ')
    
    if is_exception and 'status' not in df.columns:
        status_values = [''] + ['non-compliant'] * (len(df) - 1) if len(df) > 1 else ['']
        df['status'] = status_values

    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("Audit Snapshot")

    standard_format = workbook.add_format({'font_color': 'black'})
    header1_format = workbook.add_format({'font_color': 'black', 'bold': True})
    header2_format = workbook.add_format({'font_color': 'red', 'bold': True})
    yellow_fill = workbook.add_format({'bg_color': 'yellow', 'font_color': 'black'})
    yellow_header_format = workbook.add_format({'bg_color': 'yellow', 'font_color': 'black', 'bold': True})

    column_labels = df.columns.tolist()
    reqd_values = df.iloc[0].tolist()
    meta_count = 5 if report_type.upper() == 'ENB' else 9

    bad_columns = set()
    bad_cells = set() 
    
    for col_idx, col_name in enumerate(column_labels):
        if col_idx >= meta_count and col_name != 'status':
            baseline_str = str(reqd_values[col_idx]).strip()
            if baseline_str in ['', ' ', 'nan']:
                continue
                
            allowed_vals = [v.strip() for v in baseline_str.split(',') if v.strip()]
            for df_row_idx in range(1, len(df)):
                live_val = df.iloc[df_row_idx, col_idx]
                live_parts = [p.strip() for p in str(live_val).split('/') if p.strip()]
                if live_parts and not any(part in allowed_vals for part in live_parts):
                    bad_cells.add((df_row_idx + 1, col_idx))
                    bad_columns.add(col_idx)

    for col_idx, label in enumerate(column_labels):
        if col_idx in bad_columns:
            worksheet.write(0, col_idx, label, yellow_header_format)
        else:
            worksheet.write(0, col_idx, label, header1_format)

    worksheet.write_row(1, 0, reqd_values, header2_format)

    live_site_rows = df.iloc[1:].values.tolist()
    for row_idx, row_values in enumerate(live_site_rows, start=2):
        for col_idx, cell_value in enumerate(row_values):
            if (row_idx, col_idx) in bad_cells:
                worksheet.write(row_idx, col_idx, cell_value, yellow_fill)
            else:
                worksheet.write(row_idx, col_idx, cell_value, standard_format)
    
    workbook.close()

        
# =========================================================================
# RUN DATA INGESTION AND AUDITING ORCHESTRATOR
# =========================================================================
if __name__ == "__main__":
    subMarket = Select_subMarket()
    print(f'\n > Launching Data Preparation Engine for: {subMarket}...')

    # 1. Fetch parameters from master Excel guidelines
    reqd_parameters, group_script_name_dict, df_headers = Samsung_Parameters(tech, level)

    # 2. Extract, clean, and deduplicate vertical data files with Rules 1 & 2
    filtered_frames, common_carrierIDs = load_data_files()

    # 3. Collapse overlapping multi-value rows vectorially
    cleaned_frames, dup_carrierIDs = checking_multiple_values_in_a_row(filtered_frames, common_carrierIDs)

    # 4. Consolidate tables horizontally side-by-side
    subMarket_report_new1, filename = merging_dataframes(cleaned_frames)

    # 5. Run dynamic feature checks
    subMarket_report_new2 = ac_barring_mo_signaling_usage_feature(subMarket_report_new1)

    # 6. Save out the COMPLETE baseline sheet
    save_to_excel(filename, subMarket_report_new2)

    # 7. Execute intersection calculations across all 3 tiers
    common_items, new_dfs = get_common_enodebs_markets_and_carrierIds()
    common_enodebs = common_items[0]
    valid_markets = common_items[1]
    valid_carrierIds = common_items[2]

    # 8. Run dynamic exceptions filters
    print('\n > Launching Exception Filtering Engine across aligned reports...')
    all_non_compliant_enbs = set()

    for file_path, df in new_dfs.items():
        if 'ENB' in file_path:
            current_type = 'ENB'
            display_name = "Samsung ENB"
        else:
            current_type = 'CELL'
            display_name = "Samsung CELL " + ("Part 1" if "Part 1" in file_path else "Part 2")
            
        row_filtered_df = retain_only_noncompliant_data(df, report_type=current_type, report_name=display_name)
        final_exception_df = remove_compliant_columns(row_filtered_df, report_type=current_type)
        
        if len(final_exception_df) > 1:
            failed_enbids = final_exception_df.iloc[1:]['enbid'].astype(str).unique().tolist()
            all_non_compliant_enbs.update(failed_enbids)
        
        exception_filename = file_path.replace(' Report.xlsx', ' Exceptions.xlsx')
        save_to_excel(exception_filename, final_exception_df, report_type=current_type, is_exception=True)

    # 9. Formulate final Market KPI scorecard outputs
    data_file_timestamp = "UNKNOWN"
    if new_dfs:
        sample_df = list(new_dfs.values())[0]
        if len(sample_df) > 1:
            data_file_timestamp = str(sample_df.iloc[1]['date']).strip().upper()

    total_enbs = len(common_enodebs)
    num_non_compliant = len(all_non_compliant_enbs)
    num_compliant = total_enbs - num_non_compliant
    compliancy_rate = (num_compliant / total_enbs * 100) if total_enbs > 0 else 100.0

    print('\n#######################################################')
    print('#                 MARKET AUDIT SCORECARD              #')
    print('#######################################################')
    print(f' > Data Timestamp         : {data_file_timestamp}')
    print(f' > Total eNodeBs Audited  : {total_enbs}')
    print(f' > Non-Compliant eNodeBs  : {num_non_compliant}')
    print(f' > Compliancy Rate        : {compliancy_rate:.2f}%')
    print('#######################################################')

    kpi_txt_filename = f"{directory}/{subMarket}/Samsung_{subMarket}_CELL_Part_2_KPI_Scorecard.txt"
    with open(kpi_txt_filename, 'w', encoding='utf-8') as f:
        f.write('#######################################################\n')
        f.write('#                 MARKET AUDIT SCORECARD              #\n')
        f.write('#######################################################\n')
        f.write(f' > Data Timestamp         : {data_file_timestamp}\n')
        f.write(f' > Total eNodeBs Audited  : {total_enbs}\n')
        f.write(f' > Non-Compliant eNodeBs  : {num_non_compliant}\n')
        f.write(f' > Compliancy Rate        : {compliancy_rate:.2f}%\n')
        f.write('#######################################################\n')

    # 10 Generating Final Report
    source_files = {
        "ENB": "Raw Data/DEMO/Samsung_DEMO_ENB Exceptions.xlsx",
        "CELL_Part_1": "Raw Data/DEMO/Samsung_DEMO_CELL_Part_1 Exceptions.xlsx",
        "CELL_Part_2": "Raw Data/DEMO/Samsung_DEMO_CELL_Part_2 Exceptions.xlsx"
    }

    output_file = "Raw Data/DEMO/Samsung_DEMO_Compliance_Audit_Final.xlsx"

    new_wb = Workbook()
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)

    for new_sheet_name, source_file in source_files.items():
        source_wb = load_workbook(source_file)
        source_ws = source_wb.active

        new_ws = new_wb.create_sheet(title=new_sheet_name)

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                new_cell.value = cell.value

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)

        # Copy column widths
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width

        # Copy row heights
        for row_num, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height

        # Copy merged cells
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

        # Copy freeze panes and autofilter
        new_ws.freeze_panes = source_ws.freeze_panes
        new_ws.auto_filter.ref = source_ws.auto_filter.ref

    new_wb.save(output_file)

    print(f"Final combined workbook created: {output_file}")
    
    
    print(f"\n 🎉 End-to-End Pipeline Complete! Aligned Exception Reports compiled successfully.")
    end = time.ctime()
    print(f'Start: {start}\nEnd  : {end}')
